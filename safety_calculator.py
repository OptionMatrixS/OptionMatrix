import sys, os
_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_HERE = os.path.dirname(os.path.abspath(__file__))
for _p in [_ROOT, _HERE]:
    if _p not in sys.path:
        sys.path.insert(0, _p)

import streamlit as st
import pandas as pd
import io
from data_helpers import (
    NIFTY_STRIKES, SENSEX_STRIKES,
    get_option_price
)

_SS = st.session_state


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _nearest_strike(index: str, target: int) -> int:
    strikes = NIFTY_STRIKES if index == "NIFTY" else SENSEX_STRIKES
    return min(strikes, key=lambda x: abs(x - target))


def _spread_price(legs, strikes_per_leg: list) -> tuple:
    """
    Return (bid, ask, ltp) of the SPREAD for a given set of strikes.
    BID  = sum of all leg BIDs  (buy legs add, sell legs subtract)
    ASK  = sum of all leg ASKs
    LTP  = sum of all leg LTPs
    Assumes bid = ltp*0.998, ask = ltp*1.002 (replace with real orderbook later)
    """
    bid_total = ask_total = ltp_total = 0.0
    for leg, strike in zip(legs, strikes_per_leg):
        ltp = get_option_price(leg["index"], strike, leg["expiry"], leg["cp"])
        bid = round(ltp * 0.998, 2)
        ask = round(ltp * 1.002, 2)
        sign = 1 if leg["bs"] == "Buy" else -1
        ratio = leg["ratio"]
        bid_total += sign * bid  * ratio
        ask_total += sign * ask  * ratio
        ltp_total += sign * ltp  * ratio
    return round(bid_total, 2), round(ask_total, 2), round(ltp_total, 2)


def _build_table(legs: list, diffs: list, n_rows: int) -> tuple:
    """
    Build the safety matrix.
    diffs  : list of user-defined step per leg (one int per leg)
    n_rows : number of rows above and below zero
    Returns (df, diff_row_dict)
    """
    n_legs = len(legs)

    # DIFFERENCE header values (user inputs)
    diff_row = {"SERIES": "DIFFERENCE"}
    for i in range(n_legs):
        diff_row[f"LEG {i+1}"] = diffs[i]
    diff_row["BID"] = "SPREAD"
    diff_row["ASK"] = "SPREAD"
    diff_row["LTP"] = "SPREAD"

    rows = []
    for offset in range(-n_rows, n_rows + 1):
        row = {"SERIES": offset}
        strikes_this_row = []

        for i, leg in enumerate(legs):
            target  = leg["strike"] + offset * diffs[i]
            nearest = _nearest_strike(leg["index"], target)
            strikes_this_row.append(nearest)
            row[f"LEG {i+1}"] = nearest

        bid, ask, ltp = _spread_price(legs, strikes_this_row)
        row["BID"] = bid
        row["ASK"] = ask
        row["LTP"] = ltp

        rows.append(row)

    col_order = ["SERIES"] + [f"LEG {i+1}" for i in range(n_legs)] + ["BID", "ASK", "LTP"]
    df = pd.DataFrame(rows)[col_order]
    return df, diff_row


# ─── Render ───────────────────────────────────────────────────────────────────

def render():
    st.markdown("""
    <div style="display:flex;align-items:center;gap:12px;margin-bottom:8px;">
      <div style="font-size:20px;font-weight:600;color:#d1d4dc;">🛡️ Safety Calculator</div>
      <div style="font-size:11px;color:#787b86;padding:3px 10px;background:#1e222d;
                  border:1px solid #2a2e39;border-radius:10px;">Linked to Spread Calculator</div>
    </div>
    <div style="font-size:12px;color:#787b86;margin-bottom:16px;">
      Shows spread BID / ASK / LTP for each row of strikes.
      DIFFERENCE is user-defined per leg. Inputs sync from Spread Calculator.
    </div>
    """, unsafe_allow_html=True)

    source_legs = _SS.get("sp_legs_live", [])

    if not source_legs:
        st.info("⬅️  Go to Spread Calculator first and set up your legs, then come back here.")
        return

    n_legs = len(source_legs)

    # ── Layout ────────────────────────────────────────────────────────────────
    ctrl_col, table_col = st.columns([1, 3], gap="medium")

    with ctrl_col:
        st.markdown('<div class="sec-header">Leg Summary (from Spread Calc)</div>',
                    unsafe_allow_html=True)

        for i, leg in enumerate(source_legs):
            color = "#26a69a" if leg["bs"] == "Buy" else "#ef5350"
            st.markdown(
                f'<div style="background:#1e222d;border:1px solid #2a2e39;'
                f'border-left:3px solid {color};border-radius:5px;'
                f'padding:8px 12px;margin-bottom:6px;">'
                f'<div style="font-size:10px;color:#787b86;">LEG {i+1}</div>'
                f'<div style="font-size:12px;color:#d1d4dc;font-weight:500;">'
                f'{leg["index"]} {leg["strike"]} {leg["cp"]}</div>'
                f'<div style="font-size:11px;color:#787b86;">'
                f'{leg["expiry"]} · {leg["bs"]} · ×{leg["ratio"]}</div>'
                f'</div>',
                unsafe_allow_html=True
            )

        st.markdown("---")

        # ── FIX 2: User-defined DIFFERENCE per leg ────────────────────────────
        st.markdown('<div class="sec-header">Difference per Leg</div>',
                    unsafe_allow_html=True)
        st.markdown(
            '<div style="font-size:11px;color:#787b86;margin-bottom:8px;">'
            'Set how many points to step up/down for each leg.</div>',
            unsafe_allow_html=True
        )

        diffs = []
        for i, leg in enumerate(source_legs):
            default_diff = 50 if leg["index"] == "NIFTY" else 500
            d = st.number_input(
                f"LEG {i+1} — {leg['index']} diff",
                min_value=1,
                max_value=10000,
                value=int(_SS.get(f"sc_diff_{i}", default_diff)),
                step=default_diff,
                key=f"sc_diff_input_{i}",
                help=f"Step size for LEG {i+1} ({leg['index']}). "
                     f"Default: {'50 (Nifty)' if leg['index']=='NIFTY' else '500 (Sensex)'}"
            )
            _SS[f"sc_diff_{i}"] = d
            diffs.append(d)

        st.markdown("---")
        n_rows = st.number_input("Rows above / below", 1, 10, 5, key="sc_rows")

        calc_btn   = st.button("🛡️  Build Matrix", use_container_width=True, type="primary")
        export_btn = st.button("📥  Export Excel",  use_container_width=True)

    # ── Table ─────────────────────────────────────────────────────────────────
    with table_col:
        # Auto-rebuild if legs changed
        legs_sig = str(source_legs) + str(diffs) + str(n_rows)
        if calc_btn or _SS.get("sc_last_sig") != legs_sig:
            _SS.sc_last_sig = legs_sig
            with st.spinner("Building safety matrix..."):
                df, diff_row = _build_table(source_legs, diffs, n_rows)
                _SS.sc_df       = df
                _SS.sc_diff_row = diff_row

        if _SS.get("sc_df") is None:
            st.markdown("""
            <div style="height:300px;display:flex;align-items:center;justify-content:center;
                        background:#1e222d;border:1px solid #2a2e39;border-radius:8px;">
              <div style="text-align:center;">
                <div style="font-size:32px;margin-bottom:12px;">🛡️</div>
                <div style="font-size:14px;color:#787b86;">Click Build Matrix to generate</div>
              </div>
            </div>
            """, unsafe_allow_html=True)
            return

        df       = _SS.sc_df
        diff_row = _SS.sc_diff_row

        # ── DIFFERENCE header bar ─────────────────────────────────────────────
        diff_cells = ""
        for k, v in diff_row.items():
            if k == "SERIES":
                diff_cells += (
                    '<td style="padding:7px 10px;font-size:11px;font-weight:700;'
                    'color:#ef5350;text-align:center;min-width:60px;">DIFFERENCE</td>'
                )
            elif k.startswith("LEG"):
                diff_cells += (
                    f'<td style="padding:7px 10px;font-size:12px;font-weight:700;'
                    f'color:#ef5350;font-family:\'JetBrains Mono\',monospace;'
                    f'text-align:center;min-width:90px;">{v}</td>'
                )
            else:
                diff_cells += (
                    f'<td style="padding:7px 10px;font-size:10px;color:#787b86;'
                    f'text-align:center;min-width:75px;">{v}</td>'
                )

        # ── Column header row ─────────────────────────────────────────────────
        header_cells = (
            '<th style="padding:7px 10px;font-size:10px;color:#787b86;'
            'text-align:center;min-width:60px;">SERIES</th>'
        )
        for i in range(n_legs):
            header_cells += (
                f'<th style="padding:7px 10px;font-size:10px;color:#ef5350;'
                f'text-align:center;min-width:90px;">LEG {i+1}</th>'
            )
        for col, col_color in [("BID","#d1d4dc"), ("ASK","#d1d4dc"), ("LTP","#26a69a")]:
            header_cells += (
                f'<th style="padding:7px 10px;font-size:10px;color:{col_color};'
                f'text-align:center;min-width:75px;">{col}</th>'
            )

        # ── Data rows ─────────────────────────────────────────────────────────
        data_rows_html = ""
        for _, row in df.iterrows():
            is_zero = int(row["SERIES"]) == 0
            bg      = "#162040" if is_zero else "#1e222d"
            bl      = "border-left:3px solid #2962ff;" if is_zero else "border-left:3px solid transparent;"

            data_rows_html += f'<tr style="background:{bg};{bl}border-bottom:1px solid #2a2e39;">'

            # SERIES
            sc = "#2962ff" if is_zero else "#787b86"
            fw = "700" if is_zero else "400"
            data_rows_html += (
                f'<td style="padding:6px 10px;font-family:\'JetBrains Mono\',monospace;'
                f'font-size:12px;color:{sc};font-weight:{fw};text-align:center;">'
                f'{int(row["SERIES"])}</td>'
            )

            # LEG strikes
            for i in range(n_legs):
                vc = "#ffffff" if is_zero else "#d1d4dc"
                vw = "600" if is_zero else "400"
                data_rows_html += (
                    f'<td style="padding:6px 10px;font-family:\'JetBrains Mono\',monospace;'
                    f'font-size:12px;color:{vc};font-weight:{vw};text-align:center;">'
                    f'{int(row[f"LEG {i+1}"])}</td>'
                )

            # FIX 1: BID / ASK / LTP = SPREAD values across all legs
            for col, pos_color, neg_color in [
                ("BID", "#d1d4dc", "#ef5350"),
                ("ASK", "#d1d4dc", "#ef5350"),
                ("LTP", "#26a69a", "#ef5350"),
            ]:
                val = row[col]
                try:
                    fval = float(val)
                    color = pos_color if fval >= 0 else neg_color
                    cell_txt = f"{fval:+.2f}"
                except (ValueError, TypeError):
                    color = "#2a2e39"
                    cell_txt = "—"
                fw2 = "600" if is_zero else "400"
                data_rows_html += (
                    f'<td style="padding:6px 10px;font-family:\'JetBrains Mono\',monospace;'
                    f'font-size:12px;color:{color};font-weight:{fw2};text-align:center;">'
                    f'{cell_txt}</td>'
                )

            data_rows_html += "</tr>"

        # ── Assemble full table ───────────────────────────────────────────────
        table_html = f"""
        <div style="overflow-x:auto;border:1px solid #2a2e39;border-radius:8px;">
          <table style="width:100%;border-collapse:collapse;background:#1e222d;">
            <thead>
              <tr style="background:#1a1f2e;border-bottom:2px solid #ef535060;">
                {diff_cells}
              </tr>
              <tr style="background:#1a1f2e;border-bottom:1px solid #2a2e39;">
                {header_cells}
              </tr>
            </thead>
            <tbody>{data_rows_html}</tbody>
          </table>
        </div>
        """
        st.markdown(table_html, unsafe_allow_html=True)

        st.markdown("""
        <div style="margin-top:10px;font-size:11px;color:#787b86;display:flex;gap:20px;flex-wrap:wrap;">
          <span><span style="color:#2962ff;font-weight:700;">0</span> row = selected strikes from Spread Calculator</span>
          <span>BID / ASK / LTP = <b style="color:#d1d4dc;">spread net value</b> across all legs</span>
          <span><span style="color:#26a69a;">+</span> = net debit &nbsp; <span style="color:#ef5350;">−</span> = net credit</span>
        </div>
        """, unsafe_allow_html=True)

    # ── Excel export ──────────────────────────────────────────────────────────
    if export_btn and _SS.get("sc_df") is not None:
        df_exp   = _SS.sc_df.copy()
        diff_df  = pd.DataFrame([_SS.sc_diff_row])
        final_df = pd.concat([diff_df, df_exp], ignore_index=True)

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            final_df.to_excel(writer, index=False, sheet_name="Safety Matrix")
            wb = writer.book
            ws = writer.sheets["Safety Matrix"]

            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            thin       = Side(border_style="thin", color="2A2E39")
            border     = Border(left=thin, right=thin, top=thin, bottom=thin)
            dark_fill  = PatternFill("solid", fgColor="1E222D")
            blue_fill  = PatternFill("solid", fgColor="162040")
            hdr_fill   = PatternFill("solid", fgColor="1A1F2E")
            red_font   = Font(color="FF0000", bold=True)
            white_font = Font(color="D1D4DC")
            blue_font  = Font(color="2962FF", bold=True)

            # Row 1 = DIFFERENCE row
            for cell in ws[1]:
                cell.font      = red_font
                cell.fill      = hdr_fill
                cell.border    = border
                cell.alignment = Alignment(horizontal="center")

            # Row 2 = column headers
            for cell in ws[2]:
                cell.font      = Font(color="787B86", bold=True)
                cell.fill      = hdr_fill
                cell.border    = border
                cell.alignment = Alignment(horizontal="center")

            # Data rows start at row 3
            for row_idx in range(3, ws.max_row + 1):
                series_val = ws.cell(row=row_idx, column=1).value
                is_zero    = (series_val == 0)
                for cell in ws[row_idx]:
                    cell.border    = border
                    cell.alignment = Alignment(horizontal="center")
                    if is_zero:
                        cell.fill = blue_fill
                        cell.font = Font(color="FFFFFF", bold=True)
                    else:
                        cell.fill = dark_fill
                        cell.font = white_font

            # Auto-width
            for col in ws.columns:
                width = max(len(str(c.value or "")) for c in col) + 4
                ws.column_dimensions[col[0].column_letter].width = width

        buf.seek(0)
        st.download_button(
            label="📥 Download Safety Matrix (.xlsx)",
            data=buf,
            file_name="safety_matrix.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
