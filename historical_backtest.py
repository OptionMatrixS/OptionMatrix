import sys, os
_ROOT = os.path.dirname(os.path.abspath(__file__))
if _ROOT not in sys.path: sys.path.insert(0, _ROOT)

import streamlit as st
import plotly.graph_objects as go
import pandas as pd
import io
from datetime import datetime, date, timedelta
from fyers_client import (
    get_expiries, get_strikes, build_symbol,
    _fetch_candles, _validate_leg, RISK_FREE_RATE,
    bs_price, implied_volatility
)

_SS = st.session_state

def _init():
    for k, v in [
        ("ht_n_legs", 2), ("ht_result", None), ("ht_df", None),
        ("ht_chart_type", "Line"),
    ]:
        if k not in _SS: _SS[k] = v

def _load_expiries(index):
    cache_key = f"expiries_{index}"
    if _SS.get(cache_key): return list(_SS[cache_key].keys()), None
    try:
        exps = get_expiries(index)
        return exps, None
    except Exception as e:
        return [], str(e)

def _load_strikes(index, expiry):
    if not expiry: return [], "No expiry"
    cache_key = f"strikes_{index}_{expiry}"
    if _SS.get(cache_key): return _SS[cache_key], None
    try:
        return get_strikes(index, expiry), None
    except Exception as e:
        return [], str(e)

def _fetch_spread_history(legs, date_str, interval):
    """Fetch historical OHLCV for spread on a given date."""
    spread_close = None
    base_times   = None
    for leg in legs:
        sym = build_symbol(leg["index"], leg["expiry"], leg["cp"], leg["strike"])
        df  = _fetch_candles(sym, interval, date_str)
        if df.empty:
            raise ValueError(f"No data for {sym} on {date_str}. "
                             f"Check the date is a trading day and the contract existed.")
        price = df["close"] * leg["ratio"]
        price = price if leg["bs"] == "Buy" else -price
        if spread_close is None:
            spread_close = price
            base_times   = df.index
        else:
            spread_close = spread_close.reindex(base_times).add(
                price.reindex(base_times), fill_value=0)

    out          = pd.DataFrame(index=base_times)
    out["close"] = spread_close.values
    out["open"]  = out["close"].shift(1).fillna(out["close"])
    out["high"]  = out[["open","close"]].max(axis=1)
    out["low"]   = out[["open","close"]].min(axis=1)
    out          = out.reset_index().rename(columns={"time":"time"})
    if "index" in out.columns:
        out = out.rename(columns={"index":"time"})
    return out

def _build_chart(df, legs, chart_type, date_str, tf):
    if len(legs) >= 2:
        l = legs[0]
        title = (f"[BACKTEST {date_str}] {l['index']} {l['strike']}{l['cp']} {l['bs'][0]}"
                 f" / {legs[1]['index']} {legs[1]['strike']}{legs[1]['cp']} {legs[1]['bs'][0]}"
                 f"  [{tf}m]")
    else:
        title = f"[BACKTEST {date_str}]"

    fig = go.Figure()
    if chart_type == "Candlestick":
        fig.add_trace(go.Candlestick(
            x=df["time"], open=df["open"], high=df["high"],
            low=df["low"], close=df["close"], name="Spread",
            increasing_line_color="#26a69a", increasing_fillcolor="#26a69a",
            decreasing_line_color="#ef5350", decreasing_fillcolor="#ef5350",
            line=dict(width=1), whiskerwidth=0.3))
    else:
        fig.add_trace(go.Scatter(
            x=df["time"], y=df["close"], mode="lines", name="Spread",
            line=dict(color="#2962ff", width=1.5),
            fill="tozeroy", fillcolor="rgba(41,98,255,0.07)"))

    # Day high / low lines
    day_high = df["close"].max()
    day_low  = df["close"].min()
    fig.add_hline(y=0,        line=dict(color="#363a45", width=1, dash="dot"))
    fig.add_hline(y=day_high, line=dict(color="#26a69a", width=1, dash="dot"),
                  annotation_text=f"H: {day_high:.2f}",
                  annotation_font=dict(color="#26a69a", size=10),
                  annotation_position="right")
    fig.add_hline(y=day_low,  line=dict(color="#ef5350", width=1, dash="dot"),
                  annotation_text=f"L: {day_low:.2f}",
                  annotation_font=dict(color="#ef5350", size=10),
                  annotation_position="right")

    # Opening / closing annotations
    first_val = df["close"].iloc[0]
    last_val  = df["close"].iloc[-1]
    fig.add_annotation(x=df["time"].iloc[-1], y=last_val, text=f" {last_val:.2f}",
        showarrow=False, font=dict(size=11, color="#fff"),
        bgcolor="#26a69a" if last_val >= 0 else "#ef5350",
        borderpad=4, xanchor="left")

    fig.update_layout(
        title=dict(text=title, font=dict(size=12, color="#d1d4dc"), x=0),
        paper_bgcolor="#131722", plot_bgcolor="#131722",
        xaxis=dict(gridcolor="#1e222d", tickfont=dict(size=10, color="#787b86"),
                   rangeslider=dict(visible=False), showline=False, zeroline=False, fixedrange=False),
        yaxis=dict(gridcolor="#1e222d", tickfont=dict(size=10, color="#787b86"),
                   showline=False, zeroline=False, side="right", fixedrange=False),
        margin=dict(l=10, r=80, t=40, b=28), height=420,
        hovermode="x unified", dragmode="pan",
        hoverlabel=dict(bgcolor="#1e222d", bordercolor="#2a2e39",
                        font=dict(size=11, color="#d1d4dc")))
    return fig

def render():
    _init()
    st.markdown(
        '<div style="font-size:20px;font-weight:600;color:#d1d4dc;margin-bottom:4px;">'
        '🕰️ Historical Backtest</div>', unsafe_allow_html=True)
    st.markdown(
        '<div style="font-size:12px;color:#787b86;margin-bottom:16px;">'
        'Replay any past trading day — fetches historical intraday candles from Fyers '
        'for your spread strategy. All inputs are independent of the live Spread Chart tab.</div>',
        unsafe_allow_html=True)

    # ── Date + interval controls ──────────────────────────────────────────────
    st.markdown('<div class="sec-header">Backtest Settings</div>', unsafe_allow_html=True)
    bc1, bc2, bc3, bc4 = st.columns(4)
    with bc1:
        today     = date.today()
        # Default to last trading day
        def_date  = today - timedelta(days=1 if today.weekday() < 5 else (today.weekday() - 4))
        bt_date   = st.date_input("Backtest Date", value=def_date, key="ht_date",
                                   max_value=today - timedelta(days=1))
    with bc2:
        bt_interval = st.selectbox("Interval (min)", [1, 3, 5, 10, 15, 30, 60],
                                    index=0, key="ht_interval")
    with bc3:
        _SS.ht_chart_type = st.selectbox("Chart Type", ["Line", "Candlestick"], key="ht_ct")
    with bc4:
        _SS.ht_n_legs = st.selectbox("Legs", list(range(2, 7)),
                                      index=_SS.ht_n_legs - 2, key="ht_legs_sel")

    date_str = bt_date.strftime("%Y-%m-%d")

    # ── Leg builder ───────────────────────────────────────────────────────────
    st.markdown('<div class="sec-header">Configure Legs</div>', unsafe_allow_html=True)
    st.markdown(
        '<div style="font-size:11px;color:#787b86;margin-bottom:8px;">'
        'Select the strikes and expiries as they were on the backtest date.</div>',
        unsafe_allow_html=True)

    n    = _SS.ht_n_legs
    legs = []
    cols_legs = st.columns(n)

    for i in range(n):
        with cols_legs[i]:
            st.markdown(
                f'<div style="font-size:10px;color:#787b86;margin-bottom:6px;background:#2a2e39;'
                f'padding:2px 8px;border-radius:10px;display:inline-block;">LEG {i+1}</div>',
                unsafe_allow_html=True)

            idx = st.selectbox("Index", ["NIFTY","SENSEX","BANKNIFTY"],
                               key=f"ht_idx_{i}", label_visibility="collapsed")

            exps, exp_err = _load_expiries(idx)
            if exp_err or not exps:
                st.markdown('<div style="font-size:11px;color:#ff9800;">⏳ Loading expiries…</div>',
                            unsafe_allow_html=True)
                if st.button("🔄", key=f"ht_re_exp_{i}"):
                    _SS.pop(f"expiries_{idx}", None); st.rerun()
                legs.append(dict(index=idx,strike=0,expiry="",cp="CE",
                                 bs="Buy" if i%2==0 else "Sell",ratio=1))
                continue

            expiry = st.selectbox("Expiry", exps, key=f"ht_exp_{i}",
                                  label_visibility="collapsed")

            strikes, str_err = _load_strikes(idx, expiry)
            if str_err or not strikes:
                st.markdown('<div style="font-size:11px;color:#ff9800;">⏳ Loading strikes…</div>',
                            unsafe_allow_html=True)
                if st.button("🔄", key=f"ht_re_str_{i}"):
                    _SS.pop(f"strikes_{idx}_{expiry}", None); st.rerun()
                legs.append(dict(index=idx,strike=0,expiry=expiry,cp="CE",
                                 bs="Buy" if i%2==0 else "Sell",ratio=1))
                continue

            atm    = {"NIFTY":22800,"SENSEX":82500,"BANKNIFTY":48000}.get(idx,strikes[len(strikes)//2])
            def_s  = min(strikes, key=lambda x: abs(x - atm))
            cur    = _SS.get(f"ht_strike_{i}")
            didx   = strikes.index(cur) if cur in strikes else strikes.index(def_s)
            strike = st.selectbox("Strike", strikes, index=didx,
                                  key=f"ht_strike_{i}", label_visibility="collapsed")
            cp     = st.selectbox("CE/PE", ["CE","PE"], key=f"ht_cp_{i}",
                                  label_visibility="collapsed")
            bs     = st.selectbox("Buy/Sell", ["Buy","Sell"],
                                  index=0 if i%2==0 else 1,
                                  key=f"ht_bs_{i}", label_visibility="collapsed")
            ratio  = st.number_input("Ratio", 1, 10, 1, key=f"ht_ratio_{i}",
                                     label_visibility="collapsed")

            legs.append(dict(index=idx, strike=strike, expiry=expiry,
                             cp=cp, bs=bs, ratio=ratio))

    # ── Time range filter ─────────────────────────────────────────────────────
    st.markdown('<div class="sec-header" style="margin-top:4px;">Time Range (Optional)</div>',
                unsafe_allow_html=True)
    tr1, tr2, _ = st.columns([1, 1, 2])
    with tr1:
        time_from = st.time_input("From", value=pd.Timestamp("09:15").time(), key="ht_time_from")
    with tr2:
        time_to   = st.time_input("To",   value=pd.Timestamp("15:30").time(), key="ht_time_to")

    valid_legs = [l for l in legs if l.get("expiry") and l.get("strike",0) > 0]
    missing    = n - len(valid_legs)
    if missing:
        st.warning(f"⚠️  {missing} leg(s) still loading — wait for expiries/strikes to populate.")
    else:
        if st.button("⏮  Run Backtest", use_container_width=True, type="primary"):
            with st.spinner(f"Fetching historical data for {date_str}…"):
                try:
                    df = _fetch_spread_history(legs, date_str, bt_interval)
                    # Apply time filter
                    df["time"] = pd.to_datetime(df["time"])
                    mask = (df["time"].dt.time >= time_from) & (df["time"].dt.time <= time_to)
                    df   = df[mask].reset_index(drop=True)
                    if df.empty:
                        st.error(f"No candles in selected time range for {date_str}.")
                    else:
                        _SS.ht_df     = df
                        _SS.ht_legs   = legs
                        _SS.ht_date   = date_str
                        _SS.ht_tf     = bt_interval
                        st.rerun()
                except Exception as e:
                    st.error(f"Backtest failed: {e}")
                    st.info("Make sure the date is a trading day, the contract was active, "
                            "and your Fyers token is valid.")

    # ── Results ───────────────────────────────────────────────────────────────
    if _SS.ht_df is not None:
        df   = _SS.ht_df
        stored_legs = _SS.get("ht_legs", legs)
        dt   = _SS.get("ht_date", date_str)
        tf   = _SS.get("ht_tf", bt_interval)

        st.markdown("---")

        # Chart
        fig = _build_chart(df, stored_legs, _SS.ht_chart_type, dt, tf)
        st.plotly_chart(fig, use_container_width=True,
                        config={"displayModeBar":True,"displaylogo":False,"scrollZoom":True,
                                "toImageButtonOptions":{"format":"png","filename":f"backtest_{dt}"}})

        # Stats
        opens  = df["close"].iloc[0]
        closes = df["close"].iloc[-1]
        hi     = df["close"].max()
        lo     = df["close"].min()
        chg    = closes - opens
        chg_pct= (chg / abs(opens) * 100) if opens != 0 else 0

        st.markdown('<div class="sec-header">Day Statistics</div>', unsafe_allow_html=True)
        stats_cols = st.columns(6)
        for col, (label, val, color) in zip(stats_cols, [
            ("Open",    f"{opens:.2f}",       "#d1d4dc"),
            ("Close",   f"{closes:.2f}",      "#26a69a" if closes>=0 else "#ef5350"),
            ("High",    f"{hi:.2f}",          "#26a69a"),
            ("Low",     f"{lo:.2f}",          "#ef5350"),
            ("Change",  f"{chg:+.2f}",        "#26a69a" if chg>=0 else "#ef5350"),
            ("Chg %",   f"{chg_pct:+.2f}%",   "#26a69a" if chg_pct>=0 else "#ef5350"),
        ]):
            with col:
                st.markdown(f'<div class="stat-chip"><div class="sc-label">{label}</div>'
                            f'<div class="sc-val" style="color:{color};">{val}</div></div>',
                            unsafe_allow_html=True)

        # Data table preview
        st.markdown('<div class="sec-header" style="margin-top:12px;">Candle Data</div>',
                    unsafe_allow_html=True)
        df_show = df.copy()
        df_show["time"] = pd.to_datetime(df_show["time"]).dt.strftime("%H:%M")
        df_show = df_show.round(2)
        st.dataframe(df_show, use_container_width=True, hide_index=True, height=250)

        # ── Download ─────────────────────────────────────────────────────────
        st.markdown("---")
        dl1, dl2, _ = st.columns([1, 1, 2])

        with dl1:
            # Excel download
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                df_show.to_excel(writer, index=False, sheet_name=f"Backtest_{dt}")
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                wb = writer.book; ws = writer.sheets[f"Backtest_{dt}"]
                hfill = PatternFill("solid", fgColor="1A1F2E")
                thin  = Side(border_style="thin", color="2A2E39")
                bdr   = Border(left=thin,right=thin,top=thin,bottom=thin)
                for cell in ws[1]:
                    cell.font = Font(color="787B86", bold=True)
                    cell.fill = hfill; cell.border = bdr
                    cell.alignment = Alignment(horizontal="center")
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.fill   = PatternFill("solid", fgColor="1E222D")
                        cell.font   = Font(color="D1D4DC")
                        cell.border = bdr
                        cell.alignment = Alignment(horizontal="center")
                for col in ws.columns:
                    ws.column_dimensions[col[0].column_letter].width = 14
            buf.seek(0)
            st.download_button("📥 Download Excel",
                               data=buf,
                               file_name=f"backtest_{dt}_{stored_legs[0]['index'] if stored_legs else 'spread'}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)

        with dl2:
            csv = df_show.to_csv(index=False).encode()
            st.download_button("📄 Download CSV",
                               data=csv,
                               file_name=f"backtest_{dt}.csv",
                               mime="text/csv",
                               use_container_width=True)
