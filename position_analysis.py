import sys, os
_ROOT = os.path.dirname(os.path.abspath(__file__))
if _ROOT not in sys.path: sys.path.insert(0, _ROOT)

import streamlit as st
import pandas as pd
import io

_SS = st.session_state

DISPLAY_COLS = [
    "ID", "Underlying", "Expiry Date", "Strike Price",
    "Scrip Type", "Net Position CF", "Price CF",
    "MTM", "Net Position", "BEP", "LTP",
    "IV", "Delta", "Vega", "Gamma", "Theta",
]
NUMERIC_COLS = [
    "Net Position CF", "Price CF", "MTM", "Net Position",
    "IV", "Delta", "Vega", "Gamma", "Theta",
]
COL_RENAME = {
    "ID": "ID", "Underlying": "Underlying", "Expiry Date": "Expiry",
    "Strike Price": "Strike", "Scrip Type": "Type",
    "Net Position CF": "Net Pos CF", "Price CF": "Price CF",
    "MTM": "MTM", "Net Position": "Net Pos", "BEP": "BEP", "LTP": "LTP",
    "IV": "IV %", "Delta": "Delta", "Vega": "Vega", "Gamma": "Gamma", "Theta": "Theta",
}

def _clean_numeric(df):
    for col in NUMERIC_COLS:
        if col in df.columns:
            df[col] = (df[col].astype(str)
                       .str.replace(",", "", regex=False)
                       .str.replace(" ", "", regex=False))
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df

def _fmt(val, decimals=2):
    if pd.isna(val): return "—"
    try:
        v = float(val)
        if abs(v) >= 1_00_000: return f"{v/1_00_000:+.2f}L"
        return f"{v:+,.{decimals}f}"
    except Exception:
        return str(val)

def _color(val):
    try:
        v = float(str(val).replace(",", ""))
        if v > 0: return "#26a69a"
        if v < 0: return "#ef5350"
    except Exception:
        pass
    return "#d1d4dc"

def _init():
    for k, v in [
        ("pos_df", None), ("pos_groups", {}),
        ("pos_checked", set()), ("pos_group_checked", {}),
    ]:
        if k not in _SS: _SS[k] = v

# ─── Convert a filtered dataframe row → spread leg dict ──────────────────────
def _row_to_leg(row) -> dict:
    """Map a position row to the leg format used by spread_chart."""
    underlying = str(row.get("Underlying", "NIFTY")).strip().upper()
    # Normalise: BANKNIFTY → BANKNIFTY, SENSEX → SENSEX, else NIFTY
    if underlying not in ("NIFTY", "SENSEX", "BANKNIFTY"):
        underlying = "NIFTY"

    try:
        expiry_dt = pd.to_datetime(row.get("Expiry Date"))
        # Format as Fyers expiry label "DD Mon YY (M)" - approximate as monthly
        expiry_label = expiry_dt.strftime("%-d %b %y") + " (M)"
    except Exception:
        expiry_label = str(row.get("Expiry Date", ""))

    try:
        strike = int(float(str(row.get("Strike Price", 0)).replace(",", "")))
    except Exception:
        strike = 0

    cp    = str(row.get("Scrip Type", "CE")).strip().upper()
    if cp not in ("CE", "PE"): cp = "CE"

    # Net Position CF sign → Buy if positive, Sell if negative
    try:
        net_cf = float(str(row.get("Net Position CF", 0)).replace(",", ""))
    except Exception:
        net_cf = 0
    bs    = "Buy" if net_cf >= 0 else "Sell"
    ratio = max(1, min(10, abs(int(net_cf)) if net_cf != 0 else 1))

    try:
        ltp = float(str(row.get("LTP", 0)).replace(",", ""))
    except Exception:
        ltp = 0.0
    signed = ltp * ratio if bs == "Buy" else -ltp * ratio

    return dict(
        index=underlying, strike=strike, expiry=expiry_label,
        cp=cp, bs=bs, ratio=ratio, ltp=ltp, net=round(signed, 2)
    )

# ─── Main render ──────────────────────────────────────────────────────────────
def render():
    _init()

    st.markdown(
        '<div style="font-size:20px;font-weight:600;color:#d1d4dc;margin-bottom:4px;">'
        '📂 Position Data Analysis</div>', unsafe_allow_html=True)
    st.markdown(
        '<div style="font-size:12px;color:#787b86;margin-bottom:12px;">'
        'Upload positions · Tick rows · Send selected to Spread Chart in 1 click · '
        'Group positions for combined analysis</div>', unsafe_allow_html=True)

    # ── Upload ────────────────────────────────────────────────────────────────
    uploaded = st.file_uploader(
        "Upload position file (.xlsx or .csv)",
        type=["xlsx", "csv", "xls"], key="pos_upload")

    if uploaded is not None:
        try:
            if uploaded.name.endswith(".csv"):
                # Try multiple encodings for CSV
                raw_bytes = uploaded.read()
                for enc in ["utf-8","latin-1","cp1252","utf-8-sig","iso-8859-1"]:
                    try:
                        import io as _io
                        df_raw = pd.read_csv(_io.BytesIO(raw_bytes), encoding=enc)
                        break
                    except (UnicodeDecodeError, Exception):
                        continue
                else:
                    raise ValueError("Could not decode CSV. Try saving as UTF-8.")
            else:
                df_raw = pd.read_excel(uploaded)
            df_raw = _clean_numeric(df_raw)
            df_raw["_row_id"] = range(len(df_raw))
            _SS.pos_df = df_raw
            _SS.pos_checked = set()
            st.success(f"✅ Loaded {len(df_raw)} rows × {len(df_raw.columns)} columns")
        except Exception as e:
            st.error(f"Failed to load file: {e}"); return
    elif _SS.pos_df is not None:
        df_raw = _SS.pos_df
        if "_row_id" not in df_raw.columns:
            df_raw["_row_id"] = range(len(df_raw))
            _SS.pos_df = df_raw
        st.info("📋 Using previously uploaded file.")
    else:
        st.markdown("""
        <div style="height:180px;display:flex;align-items:center;justify-content:center;
                    background:#1e222d;border:1px dashed #2a2e39;border-radius:8px;">
          <div style="text-align:center;">
            <div style="font-size:32px;margin-bottom:8px;">📂</div>
            <div style="font-size:13px;color:#787b86;">Upload a position file to begin</div>
          </div>
        </div>""", unsafe_allow_html=True)
        return

    df_raw = _SS.pos_df

    # ── Filters ───────────────────────────────────────────────────────────────
    st.markdown('<div class="sec-header">Filters</div>', unsafe_allow_html=True)
    fc1, fc2, fc3, fc4, fc5 = st.columns(5)

    with fc1:
        ids = sorted(df_raw["ID"].dropna().unique().tolist()) if "ID" in df_raw.columns else []
        sel_ids = st.multiselect("ID", ids, default=ids[:1] if ids else [], key="pos_ids")

    with fc2:
        unds = sorted(df_raw["Underlying"].dropna().unique().tolist()) if "Underlying" in df_raw.columns else []
        sel_und = st.multiselect("Underlying", unds, default=unds, key="pos_und")

    with fc3:
        types = sorted(df_raw["Scrip Type"].dropna().unique().tolist()) if "Scrip Type" in df_raw.columns else []
        sel_type = st.multiselect("Option Type", types, default=types, key="pos_type")

    with fc4:
        sel_expiries = []
        if "Expiry Date" in df_raw.columns:
            try:
                df_raw["Expiry Date"] = pd.to_datetime(df_raw["Expiry Date"], errors="coerce")
                expiries   = sorted(df_raw["Expiry Date"].dropna().unique())
                exp_labels = [pd.Timestamp(e).strftime("%d %b %y") for e in expiries]
                sel_exp_l  = st.multiselect("Expiry", exp_labels, default=exp_labels, key="pos_exp")
                sel_expiries = [expiries[exp_labels.index(l)] for l in sel_exp_l if l in exp_labels]
            except Exception:
                pass

    with fc5:
        show_greeks = st.checkbox("Show Greeks", value=False, key="pos_greeks")

    # Strike filter — shown after main row
    strike_filter_vals = []
    if "Strike Price" in df_raw.columns:
        df_for_strikes = df_raw.copy()
        if sel_ids and "ID" in df_for_strikes.columns:
            df_for_strikes = df_for_strikes[df_for_strikes["ID"].isin(sel_ids)]
        avail_strikes = sorted([int(x) for x in df_for_strikes["Strike Price"].dropna().unique() if x > 0])
        if avail_strikes:
            strike_filter_vals = st.multiselect(
                "Strike Price", avail_strikes, default=avail_strikes, key="pos_strikes",
                help="Filter by specific strike prices")

    # ── Apply filters ─────────────────────────────────────────────────────────
    df = df_raw.copy()
    if sel_ids    and "ID"          in df.columns: df = df[df["ID"].isin(sel_ids)]
    if sel_und    and "Underlying"  in df.columns: df = df[df["Underlying"].isin(sel_und)]
    if sel_type   and "Scrip Type"  in df.columns: df = df[df["Scrip Type"].isin(sel_type)]
    if sel_expiries and "Expiry Date" in df.columns: df = df[df["Expiry Date"].isin(sel_expiries)]
    if strike_filter_vals and "Strike Price" in df.columns:
        df = df[df["Strike Price"].isin(strike_filter_vals)]

    sort_cols = [c for c in ["Underlying", "Expiry Date", "Strike Price"] if c in df.columns]
    if sort_cols: df = df.sort_values(sort_cols)
    if df.empty: st.warning("No rows match the selected filters."); return

    # ── Summary bar ───────────────────────────────────────────────────────────
    def safe_sum(col): return df[col].sum() if col in df.columns else 0
    mtm = safe_sum("MTM"); delta = safe_sum("Delta"); vega = safe_sum("Vega")
    m1,m2,m3,m4,m5 = st.columns(5)
    for col,(lbl,val,clr) in zip([m1,m2,m3,m4,m5],[
        ("Total MTM",    _fmt(mtm),   "#26a69a" if mtm>=0 else "#ef5350"),
        ("Net Position", _fmt(safe_sum("Net Position")), "#d1d4dc"),
        ("Net Delta",    f"{delta:+.2f}", "#2962ff"),
        ("Net Vega",     f"{vega:+.2f}",  "#9c27b0"),
        ("Rows",         str(len(df)),    "#787b86"),
    ]):
        with col:
            st.markdown(f'<div class="stat-chip"><div class="sc-label">{lbl}</div>'
                        f'<div class="sc-val" style="color:{clr};">{val}</div></div>',
                        unsafe_allow_html=True)

    st.markdown('<div style="height:6px"></div>', unsafe_allow_html=True)

    # ── SELECT-ALL bar ────────────────────────────────────────────────────────
    st.markdown('<div class="sec-header">Position Table  <span style="font-weight:400;color:#2a2e39;">— tick rows, then send to Spread Chart</span></div>',
                unsafe_allow_html=True)

    sa_col, _, act_col = st.columns([1, 2, 3])
    with sa_col:
        if st.button("☑ Select All", key="pos_sel_all", use_container_width=True):
            _SS.pos_checked = set(df["_row_id"].tolist())
            st.rerun()
    with act_col:
        if st.button("☐ Clear Selection", key="pos_clear_sel", use_container_width=True):
            _SS.pos_checked = set()
            st.rerun()

    # ── Position table with checkboxes ────────────────────────────────────────
    cols_show = [c for c in DISPLAY_COLS if c in df.columns]
    if not show_greeks:
        cols_show = [c for c in cols_show if c not in ("IV","Delta","Vega","Gamma","Theta")]

    # Build rows using st.columns for checkbox + data side by side
    # Header
    header_cols = st.columns([0.3] + [1]*len(cols_show))
    header_cols[0].markdown(
        '<div style="font-size:10px;color:#787b86;padding:4px 0;text-align:center;">✓</div>',
        unsafe_allow_html=True)
    for ci, col_name in enumerate(cols_show):
        header_cols[ci+1].markdown(
            f'<div style="font-size:10px;color:#787b86;padding:4px 2px;'
            f'border-bottom:1px solid #2a2e39;white-space:nowrap;">'
            f'{COL_RENAME.get(col_name, col_name)}</div>',
            unsafe_allow_html=True)

    st.markdown('<div style="height:2px;background:#2a2e39;margin-bottom:2px;"></div>',
                unsafe_allow_html=True)

    for _, row in df.iterrows():
        row_id  = int(row["_row_id"])
        checked = row_id in _SS.pos_checked
        is_ce   = str(row.get("Scrip Type","")).upper() == "CE"

        row_cols = st.columns([0.3] + [1]*len(cols_show))

        # Checkbox
        new_val = row_cols[0].checkbox(
            "", value=checked, key=f"pos_chk_{row_id}",
            label_visibility="collapsed")
        if new_val != checked:
            if new_val: _SS.pos_checked.add(row_id)
            else:       _SS.pos_checked.discard(row_id)

        # Row bg colour via markdown container
        bg = "#1a2744" if row_id in _SS.pos_checked else "#1e222d"

        for ci, col_name in enumerate(cols_show):
            val = row.get(col_name, "")
            cell_md = ""
            if col_name == "Scrip Type":
                clr = "#2962ff" if is_ce else "#ff9800"
                cell_md = (f'<div style="font-size:11px;color:{clr};font-weight:600;'
                           f'background:{bg};padding:4px 2px;border-bottom:1px solid #2a2e39;">{val}</div>')
            elif col_name in NUMERIC_COLS:
                clr = _color(val)
                cell_md = (f'<div style="font-size:11px;color:{clr};background:{bg};'
                           f'padding:4px 2px;border-bottom:1px solid #2a2e39;'
                           f'font-family:\'JetBrains Mono\',monospace;text-align:right;">'
                           f'{_fmt(val)}</div>')
            elif col_name == "Expiry Date":
                try:    v = pd.to_datetime(val).strftime("%d %b %y")
                except: v = str(val)
                cell_md = (f'<div style="font-size:11px;color:#d1d4dc;background:{bg};'
                           f'padding:4px 2px;border-bottom:1px solid #2a2e39;text-align:center;">{v}</div>')
            elif col_name == "Strike Price":
                try:    sv = str(int(float(val)))
                except: sv = str(val)
                cell_md = (f'<div style="font-size:11px;color:#d1d4dc;font-weight:500;background:{bg};'
                           f'padding:4px 2px;border-bottom:1px solid #2a2e39;'
                           f'font-family:\'JetBrains Mono\',monospace;text-align:center;">{sv}</div>')
            else:
                cell_md = (f'<div style="font-size:11px;color:#d1d4dc;background:{bg};'
                           f'padding:4px 2px;border-bottom:1px solid #2a2e39;">{val}</div>')

            row_cols[ci+1].markdown(cell_md, unsafe_allow_html=True)

    # ── Totals row with BEP ──────────────────────────────────────────────────
    try:
        net_cf_total  = df["Net Position CF"].sum()  if "Net Position CF"  in df.columns else 0
        net_pos_total = df["Net Position"].sum()      if "Net Position"     in df.columns else 0

        if "BEP" in df.columns and "Net Position" in df.columns:
            weights = df["Net Position"].abs()
            w_sum   = weights.sum()
            bep_total = (df["BEP"] * weights).sum() / w_sum if w_sum > 0 else 0
        else:
            bep_total = 0

        # Sign rule: follow Net Position; if 0, follow Net Position CF
        ref_sign = net_pos_total if net_pos_total != 0 else net_cf_total
        if ref_sign > 0:
            bep_total = abs(bep_total)
        elif ref_sign < 0:
            bep_total = -abs(bep_total)

        qty_total   = abs(net_pos_total) if net_pos_total != 0 else 1
        bep_per_qty = round(bep_total / qty_total, 2) if qty_total else 0
    except Exception:
        bep_total = bep_per_qty = 0

    total_cols = st.columns([0.3] + [1]*len(cols_show))
    total_cols[0].markdown(
        f'<div style="font-size:10px;color:#2962ff;padding:5px 0;'
        f'border-top:2px solid #2962ff;text-align:center;">'
        f'{len(df)}</div>', unsafe_allow_html=True)
    for ci, col_name in enumerate(cols_show):
        if col_name == "ID":
            total_cols[ci+1].markdown(
                '<div style="font-size:11px;font-weight:700;color:#2962ff;'
                'border-top:2px solid #2962ff;padding:5px 2px;">TOTAL</div>',
                unsafe_allow_html=True)
        elif col_name == "BEP":
            clr = _color(bep_total)
            total_cols[ci+1].markdown(
                f'<div style="font-size:11px;font-weight:700;color:{clr};'
                f'border-top:2px solid #2962ff;padding:5px 2px;'
                f'font-family:\'JetBrains Mono\',monospace;text-align:right;">'
                f'{_fmt(bep_total)}'
                f'<br><span style="font-size:9px;color:#787b86;">÷qty: {_fmt(bep_per_qty)}</span></div>',
                unsafe_allow_html=True)
        elif col_name in NUMERIC_COLS:
            tot = df[col_name].sum() if col_name in df.columns else 0
            clr = _color(tot)
            total_cols[ci+1].markdown(
                f'<div style="font-size:11px;font-weight:700;color:{clr};'
                f'border-top:2px solid #2962ff;padding:5px 2px;'
                f'font-family:\'JetBrains Mono\',monospace;text-align:right;">'
                f'{_fmt(tot)}</div>', unsafe_allow_html=True)
        else:
            total_cols[ci+1].markdown(
                '<div style="border-top:2px solid #2962ff;padding:5px 2px;"></div>',
                unsafe_allow_html=True)

    # ── SEND TO SPREAD CHART ──────────────────────────────────────────────────
    n_checked = len(_SS.pos_checked)
    st.markdown("---")

    sc1, sc2, sc3 = st.columns([2, 2, 2])

    with sc1:
        # Show how many are ticked
        checked_color = "#26a69a" if n_checked > 0 else "#787b86"
        st.markdown(
            f'<div style="background:#1e222d;border:1px solid #2a2e39;border-radius:6px;'
            f'padding:10px 14px;margin-bottom:8px;">'
            f'<div style="font-size:11px;color:#787b86;">Selected rows</div>'
            f'<div style="font-size:22px;font-weight:600;color:{checked_color};">'
            f'{n_checked}</div>'
            f'<div style="font-size:10px;color:#787b86;">of {len(df)} filtered rows</div>'
            f'</div>', unsafe_allow_html=True)

    with sc2:
        st.markdown('<div style="height:4px"></div>', unsafe_allow_html=True)
        btn_disabled = n_checked == 0 or n_checked > 6
        if n_checked > 6:
            st.warning(f"⚠️ Max 6 legs in Spread Chart. You selected {n_checked}. Tick ≤ 6 rows.")
        elif n_checked == 0:
            st.info("Tick rows above to select positions.")

        sb1, sb2 = st.columns(2)
        with sb1:
         if st.button(
            f"📊  Check {n_checked} in Spread Chart",
            use_container_width=True,
            type="primary",
            disabled=btn_disabled,
            key="pos_send_to_spread"
         ):
            # Build legs from checked rows
            checked_rows = df[df["_row_id"].isin(_SS.pos_checked)]
            legs = []
            for _, row in checked_rows.iterrows():
                try:
                    leg = _row_to_leg(row)
                    legs.append(leg)
                except Exception:
                    pass

            if legs:
                # Inject into spread chart session state
                n = len(legs)
                _SS.sp_n_legs   = n
                _SS.sp_legs_live = legs

                # Set individual widget keys so spread chart reads them correctly
                for i, leg in enumerate(legs):
                    _SS[f"sp_idx_{i}"]    = leg["index"]
                    _SS[f"sp_strike_{i}"] = leg["strike"]
                    _SS[f"sp_exp_{i}"]    = leg["expiry"]
                    _SS[f"sp_cp_{i}"]     = leg["cp"]
                    _SS[f"sp_bs_{i}"]     = leg["bs"]
                    _SS[f"sp_ratio_{i}"]  = leg["ratio"]

                # Clear old result so chart reloads fresh
                _SS.sp_result = None
                _SS.sp_df     = None

                # Navigate to spread chart
                _SS.page = "spread"
                st.success(f"✅ {n} legs loaded into Spread Chart!")
                st.rerun()
        with sb2:
         if st.button(
            f"🏗️  Send {n_checked} to Strategy Builder",
            use_container_width=True,
            type="secondary",
            disabled=btn_disabled,
            key="pos_send_to_sb"
         ):
            checked_rows2 = df[df["_row_id"].isin(_SS.pos_checked)]
            sb_legs = []
            for _, row2 in checked_rows2.head(10).iterrows():
                try: sb_legs.append(_row_to_leg(row2))
                except Exception: pass
            if sb_legs:
                for i2, leg2 in enumerate(sb_legs):
                    _SS[f"sb_bs_{i2}"]     = leg2["bs"]
                    _SS[f"sb_cp_{i2}"]     = leg2["cp"]
                    _SS[f"sb_strike_{i2}"] = leg2["strike"]
                    _SS[f"sb_prem_{i2}"]   = leg2["ltp"]
                _SS.sb_n_legs = len(sb_legs)
                _SS.page      = "strategy"
                st.rerun()

    with sc3:
        # Show selected legs preview
        if n_checked > 0:
            checked_rows = df[df["_row_id"].isin(_SS.pos_checked)]
            preview_html = ""
            for _, row in checked_rows.head(6).iterrows():
                und  = row.get("Underlying","")
                stk  = row.get("Strike Price","")
                cp   = row.get("Scrip Type","")
                exp  = ""
                try: exp = pd.to_datetime(row.get("Expiry Date")).strftime("%d %b")
                except: pass
                clr = "#2962ff" if str(cp).upper()=="CE" else "#ff9800"
                preview_html += (
                    f'<div style="font-size:11px;color:#d1d4dc;padding:2px 0;">'
                    f'{und} <b style="color:{clr};">{int(stk) if not pd.isna(stk) else stk}</b> '
                    f'<span style="color:{clr};">{cp}</span> {exp}</div>'
                )
            st.markdown(
                f'<div style="background:#1e222d;border:1px solid #2a2e39;border-radius:6px;'
                f'padding:10px 14px;">'
                f'<div style="font-size:10px;color:#787b86;margin-bottom:6px;">Selected positions</div>'
                f'{preview_html}</div>',
                unsafe_allow_html=True)

    # ── GROUPS ────────────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown('<div class="sec-header">Position Groups</div>', unsafe_allow_html=True)
    st.markdown(
        '<div style="font-size:11px;color:#787b86;margin-bottom:10px;">'
        'Create named groups of positions for combined MTM / Greeks analysis.</div>',
        unsafe_allow_html=True)

    # Create new group
    gc1, gc2 = st.columns([2, 1])
    with gc1:
        new_group_name = st.text_input(
            "Group name", placeholder="e.g. BankNifty Jun Spreads",
            key="pos_new_group_name", label_visibility="collapsed")
    with gc2:
        if st.button("➕ Create group from selection", use_container_width=True,
                     disabled=(n_checked == 0 or not new_group_name),
                     key="pos_create_group"):
            if new_group_name and n_checked > 0:
                if "pos_groups" not in _SS: _SS.pos_groups = {}
                _SS.pos_groups[new_group_name] = list(_SS.pos_checked)
                st.success(f"Group '{new_group_name}' created with {n_checked} positions.")
                st.rerun()

    # Show existing groups
    if _SS.pos_groups:
        for grp_name, row_ids in list(_SS.pos_groups.items()):
            grp_df = df_raw[df_raw["_row_id"].isin(row_ids)].copy()
            if grp_df.empty:
                del _SS.pos_groups[grp_name]
                continue

            grp_mtm   = grp_df["MTM"].sum()   if "MTM"   in grp_df.columns else 0
            grp_delta = grp_df["Delta"].sum()  if "Delta" in grp_df.columns else 0
            grp_vega  = grp_df["Vega"].sum()   if "Vega"  in grp_df.columns else 0
            grp_color = "#26a69a" if grp_mtm >= 0 else "#ef5350"

            with st.expander(
                f"📁  {grp_name}  ·  {len(grp_df)} positions  ·  "
                f"MTM: {_fmt(grp_mtm)}", expanded=False
            ):
                # Group summary chips
                gm1, gm2, gm3, gm4 = st.columns(4)
                for gcol, (lbl, val, clr) in zip(
                    [gm1, gm2, gm3, gm4],
                    [
                        ("Group MTM",    _fmt(grp_mtm),   grp_color),
                        ("Net Delta",    f"{grp_delta:+.2f}", "#2962ff"),
                        ("Net Vega",     f"{grp_vega:+.2f}",  "#9c27b0"),
                        ("Positions",    str(len(grp_df)),    "#787b86"),
                    ]
                ):
                    with gcol:
                        st.markdown(
                            f'<div class="stat-chip">'
                            f'<div class="sc-label">{lbl}</div>'
                            f'<div class="sc-val" style="color:{clr};">{val}</div>'
                            f'</div>', unsafe_allow_html=True)

                st.markdown('<div style="height:6px"></div>', unsafe_allow_html=True)

                # Show mini table
                show_cols = [c for c in ["Underlying","Expiry Date","Strike Price",
                                          "Scrip Type","Net Position CF","MTM","LTP"]
                             if c in grp_df.columns]
                grp_display = grp_df[show_cols].copy()
                if "Expiry Date" in grp_display.columns:
                    grp_display["Expiry Date"] = pd.to_datetime(
                        grp_display["Expiry Date"], errors="coerce"
                    ).dt.strftime("%d %b %y")
                st.dataframe(grp_display, use_container_width=True, hide_index=True, height=200)

                # Group action buttons
                ba1, ba2, ba3 = st.columns(3)
                with ba1:
                    if st.button(f"📊 Send to Spread Chart",
                                 key=f"grp_send_{grp_name}", use_container_width=True,
                                 disabled=len(grp_df) > 6):
                        legs = []
                        for _, row in grp_df.head(6).iterrows():
                            try: legs.append(_row_to_leg(row))
                            except Exception: pass
                        if legs:
                            n = len(legs)
                            _SS.sp_n_legs    = n
                            _SS.sp_legs_live = legs
                            for i, leg in enumerate(legs):
                                _SS[f"sp_idx_{i}"]    = leg["index"]
                                _SS[f"sp_strike_{i}"] = leg["strike"]
                                _SS[f"sp_exp_{i}"]    = leg["expiry"]
                                _SS[f"sp_cp_{i}"]     = leg["cp"]
                                _SS[f"sp_bs_{i}"]     = leg["bs"]
                                _SS[f"sp_ratio_{i}"]  = leg["ratio"]
                            _SS.sp_result = None
                            _SS.sp_df     = None
                            _SS.page      = "spread"
                            st.rerun()
                        if len(grp_df) > 6:
                            st.caption("Max 6 legs — group has too many positions.")

                with ba2:
                    # Add current selection to group
                    if st.button(f"➕ Add selection to group",
                                 key=f"grp_add_{grp_name}", use_container_width=True,
                                 disabled=n_checked==0):
                        _SS.pos_groups[grp_name] = list(set(row_ids) | _SS.pos_checked)
                        st.rerun()

                with ba3:
                    if st.button(f"🗑️ Delete group",
                                 key=f"grp_del_{grp_name}", use_container_width=True):
                        del _SS.pos_groups[grp_name]
                        st.rerun()

    elif not _SS.pos_groups:
        st.markdown(
            '<div style="padding:16px;background:#1e222d;border:1px dashed #2a2e39;'
            'border-radius:6px;font-size:12px;color:#787b86;text-align:center;">'
            'No groups yet. Tick rows in the table above, enter a name, '
            'and click Create group.</div>', unsafe_allow_html=True)

    # ── Export ────────────────────────────────────────────────────────────────
    st.markdown("---")
    ec1, ec2 = st.columns(2)
    export_cols = [c for c in DISPLAY_COLS if c in df.columns]
    if not show_greeks:
        export_cols = [c for c in export_cols if c not in ("IV","Delta","Vega","Gamma","Theta")]
    df_export = df[export_cols].copy()

    with ec1:
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df_export.to_excel(writer, index=False, sheet_name="Positions")
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            ws  = writer.sheets["Positions"]
            hf  = PatternFill("solid", fgColor="1A1F2E")
            thin= Side(border_style="thin", color="2A2E39")
            bdr = Border(left=thin,right=thin,top=thin,bottom=thin)
            for cell in ws[1]:
                cell.font=Font(color="787B86",bold=True); cell.fill=hf
                cell.border=bdr; cell.alignment=Alignment(horizontal="center")
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.fill=PatternFill("solid",fgColor="1E222D")
                    cell.font=Font(color="D1D4DC"); cell.border=bdr
                    cell.alignment=Alignment(horizontal="center")
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(
                    max(len(str(c.value or "")) for c in col)+4, 20)
        buf.seek(0)
        st.download_button("📥 Export Filtered (Excel)", data=buf,
                           file_name="positions_filtered.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)
    with ec2:
        st.download_button("📄 Export Filtered (CSV)",
                           data=df_export.to_csv(index=False).encode(),
                           file_name="positions_filtered.csv",
                           mime="text/csv", use_container_width=True)
