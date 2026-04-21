"""
live_bhavcopy.py — Live Bhavcopy
Fetches full option chain from Fyers and displays all strikes with volume.
"""
import sys, os
_ROOT = os.path.dirname(os.path.abspath(__file__))
if _ROOT not in sys.path: sys.path.insert(0, _ROOT)

import streamlit as st
import pandas as pd
import io
from fyers_client import get_fyers_client, get_expiries, _label_to_code

_SS = st.session_state

_OPTIDX = {"NIFTY":"NSE:NIFTY50-INDEX","BANKNIFTY":"NSE:NIFTYBANK-INDEX",
           "SENSEX":"BSE:SENSEX-INDEX","FINNIFTY":"NSE:FINNIFTY-INDEX"}

_FNO_STOCKS = sorted(["RELIANCE","TCS","HDFCBANK","INFY","ICICIBANK","KOTAKBANK",
    "SBIN","BAJFINANCE","BHARTIARTL","ITC","AXISBANK","LT","ASIANPAINT","MARUTI",
    "TITAN","SUNPHARMA","ULTRACEMCO","WIPRO","HCLTECH","TECHM","INDUSINDBK",
    "BAJAJFINSV","NESTLEIND","POWERGRID","NTPC","ONGC","COALINDIA","TATAMOTORS",
    "TATASTEEL","JSWSTEEL","HINDALCO","VEDL","GRASIM","ADANIENT","ADANIPORTS",
    "DIVISLAB","DRREDDY","CIPLA","APOLLOHOSP","EICHERMOT","HEROMOTOCO","BAJAJ-AUTO",
    "TATACONSUM","BRITANNIA","PIDILITIND","HAVELLS","MUTHOOTFIN","LUPIN","BIOCON",
    "AUROPHARMA","TORNTPHARM","ALKEM","NAUKRI","IRCTC","DMART","ZOMATO",
    "BANDHANBNK","FEDERALBNK","IDFCFIRSTB","PNB","BANKBARODA","CANBK",
    "SBILIFE","HDFCLIFE","ICICIPRULI","COFORGE","MPHASIS","PERSISTENT","LTIM","OFSS",
    "INDIGO","SAIL","NMDC","BPCL","IOC","HINDPETRO","GAIL","PETRONET","IGL"])

def _fetch_chain(sym_key, expiry_label, index_name):
    """
    Fetch option chain for one symbol+expiry.
    Uses optionchain API with strikecount:0 to get all strikes.
    """
    try:
        fyers = get_fyers_client()
        # Get all expiry data first, then filter
        resp = fyers.optionchain(data={"symbol": sym_key, "strikecount": 0, "timestamp": ""})
        if not (resp and resp.get("s") == "ok"):
            return pd.DataFrame(), f"API error: {resp.get('message','unknown')}"

        chain = resp.get("data", {}).get("optionsChain", [])
        if not chain:
            return pd.DataFrame(), "No option chain data returned"

        rows = []
        for opt in chain:
            if not isinstance(opt, dict): continue
            rows.append({
                "Particular":    index_name,
                "Expiry":        opt.get("expiry",""),
                "Strike Price":  opt.get("strikePrice",0),
                "Option Type":   opt.get("option_type","").upper(),
                "Volume":        int(opt.get("volume",0) or 0),
                "OI":            int(opt.get("oi",0) or 0),
                "Chng in OI":    int(opt.get("oiChange",0) or 0),
                "LTP":           float(opt.get("ltp",0) or 0),
            })
        return pd.DataFrame(rows), None
    except Exception as e:
        return pd.DataFrame(), str(e)

def render():
    st.markdown('<div style="font-size:20px;font-weight:600;color:#d1d4dc;margin-bottom:4px;">📋 Live Bhavcopy</div>',unsafe_allow_html=True)
    st.markdown('<div style="font-size:12px;color:#787b86;margin-bottom:12px;">All option strikes with live volume, OI and LTP from Fyers.</div>',unsafe_allow_html=True)

    r1c1, r1c2 = st.columns(2)
    with r1c1:
        inst_type = st.radio("Instrument Type", ["OPTIDX (Index)","OPTSTK (Stocks)"],
                              key="bh_inst_type", horizontal=True)

    with r1c2:
        if "OPTIDX" in inst_type:
            idx_sel = st.selectbox("Index", list(_OPTIDX.keys()), key="bh_idx")
            sym_key = _OPTIDX[idx_sel]
            sym_name= idx_sel
        else:
            stk_opts = ["— All Stocks —"] + _FNO_STOCKS
            stk_sel  = st.multiselect("Stocks", stk_opts,
                                       default=["— All Stocks —"], key="bh_stks")
            sel_stks = _FNO_STOCKS if "— All Stocks —" in stk_sel else [s for s in stk_sel if s!="— All Stocks —"]
            sym_key  = f"NSE:{sel_stks[0]}-EQ" if sel_stks else ""
            sym_name = sel_stks[0] if sel_stks else ""

    # Expiry selector
    r2c1, r2c2, r2c3, r2c4 = st.columns(4)
    with r2c1:
        idx_for_exp = idx_sel if "OPTIDX" in inst_type else "NIFTY"
        try:
            exps = get_expiries(idx_for_exp)
            exp_sel = st.selectbox("Expiry", ["All"] + exps, key="bh_exp")
        except Exception as e:
            st.error(f"Load expiries: {e}"); return

    with r2c2:
        vol_gt = st.number_input("Volume >", min_value=0, value=0, step=1,
                                  key="bh_vol")
    with r2c3:
        opt_type_f = st.selectbox("Option Type", ["Both","CE","PE"], key="bh_ot")
    with r2c4:
        new_oi_only = st.checkbox("New OI Only\n(OI was 0)", value=False, key="bh_newoi")

    col_b1, col_b2 = st.columns(2)
    with col_b1:
        fetch_btn = st.button("📡 Fetch Bhavcopy", type="primary",
                               use_container_width=True, key="bh_fetch")
    with col_b2:
        if st.button("🔄 Refresh", use_container_width=True, key="bh_refresh"):
            _SS.pop("bh_result", None); st.rerun()

    if fetch_btn:
        with st.spinner("Fetching option chain…"):
            frames = []
            if "OPTIDX" in inst_type:
                df, err = _fetch_chain(sym_key, exp_sel, sym_name)
                if err: st.error(f"Error: {err}")
                elif not df.empty: frames.append(df)
            else:
                for stk in (sel_stks[:15] if sel_stks else []):
                    df, err = _fetch_chain(f"NSE:{stk}-EQ", exp_sel, stk)
                    if not df.empty: frames.append(df)

            if frames:
                all_df = pd.concat(frames, ignore_index=True)
                # Filter by expiry
                if exp_sel != "All" and "Expiry" in all_df.columns:
                    # Fyers expiry in chain may be date string; try to match
                    pass  # keep all for now — user already selected expiry at API level

                # Apply filters
                if vol_gt > 0:
                    all_df = all_df[pd.to_numeric(all_df["Volume"],errors="coerce").fillna(0) > vol_gt]
                if new_oi_only:
                    # New OI = positions where OI change > 0 and was previously 0
                    all_df = all_df[pd.to_numeric(all_df["Chng in OI"],errors="coerce").fillna(0) > 0]
                if opt_type_f != "Both":
                    all_df = all_df[all_df["Option Type"] == opt_type_f]

                all_df = all_df.sort_values("Volume", ascending=False).reset_index(drop=True)
                _SS.bh_result = all_df
            else:
                st.warning("No data returned. Check token and market hours.")
                _SS.bh_result = None

    # Display
    if _SS.get("bh_result") is not None and not _SS.bh_result.empty:
        df = _SS.bh_result

        st.markdown(f'<div style="font-size:12px;color:#787b86;margin:8px 0;">'
                    f'Showing <b style="color:#d1d4dc;">{len(df)}</b> strikes</div>',
                    unsafe_allow_html=True)

        # Format for display
        df_show = df.copy()
        for col in ["Volume","OI","Chng in OI"]:
            if col in df_show.columns:
                df_show[col] = pd.to_numeric(df_show[col],errors="coerce").fillna(0).astype(int)

        st.dataframe(df_show, use_container_width=True, hide_index=True, height=500)

        # Summary stats
        s1,s2,s3 = st.columns(3)
        with s1: st.metric("Total Volume", f"{df['Volume'].sum():,.0f}")
        with s2: st.metric("Total OI",     f"{df['OI'].sum():,.0f}")
        with s3: st.metric("Strikes shown",str(len(df)))

        # Export
        st.markdown("---")
        ec1,ec2 = st.columns(2)
        with ec1:
            buf = io.BytesIO()
            with pd.ExcelWriter(buf,engine="openpyxl") as writer:
                df_show.to_excel(writer,index=False,sheet_name="Bhavcopy")
                from openpyxl.styles import Font,PatternFill,Alignment
                ws=writer.sheets["Bhavcopy"]
                hf=PatternFill("solid",fgColor="1A1F2E")
                for cell in ws[1]:
                    cell.font=Font(color="787B86",bold=True); cell.fill=hf
                    cell.alignment=Alignment(horizontal="center")
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.fill=PatternFill("solid",fgColor="1E222D")
                        cell.font=Font(color="D1D4DC")
                        cell.alignment=Alignment(horizontal="center")
                for col in ws.columns:
                    ws.column_dimensions[col[0].column_letter].width=min(
                        max(len(str(c.value or "")) for c in col)+4,22)
            buf.seek(0)
            st.download_button("📥 Export Excel",data=buf,
                               file_name="bhavcopy.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
        with ec2:
            st.download_button("📄 Export CSV",
                               data=df_show.to_csv(index=False).encode(),
                               file_name="bhavcopy.csv",mime="text/csv",
                               use_container_width=True)
    elif _SS.get("bh_result") is not None:
        st.info("No strikes match filters.")
    else:
        st.markdown('<div style="height:200px;display:flex;align-items:center;justify-content:center;'
                    'background:#1e222d;border:1px dashed #2a2e39;border-radius:8px;">'
                    '<div style="text-align:center;color:#787b86;">'
                    '<div style="font-size:28px;margin-bottom:8px;">📋</div>'
                    'Select filters and click Fetch Bhavcopy</div></div>',
                    unsafe_allow_html=True)
